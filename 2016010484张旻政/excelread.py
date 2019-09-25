import openpyxl as xl

wb = xl.load_workbook("data.xlsx")          #加载一个excel表文件并存储在wb变量内
print(wb.sheetnames)                        #打印sheet页的名字
sheet = wb["test"]                          #将名为test的sheet储存在变量sheet内
cell = sheet['B3']                          #读取名为“test”sheet中B3单元格
print(cell.value)                           #打印该单元格的值
print(sheet.cell(row=1,column=1).value)     #打印第一行一列单元格的值
wb.close()                                  #关闭文件
